"""生成客户表.xlsx模板。用法: python build_table.py [输出路径]
表不存在时用本脚本创建；改结构时参照本脚本"整体重建"，不要insert_cols。"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = sys.argv[1] if len(sys.argv) > 1 else "客户表.xlsx"

GROUPS = [
    ("管理", ["序号", "状态", "录入日期"], "D9D9D9"),
    ("① 名片信息", ["姓名", "职位", "公司名", "邮箱", "电话", "网址",
                    "国家/地址", "展会备注", "名片照片"], "DDEBF7"),
    ("② 客户背调", ["公司简介", "客户类型", "采购角色判断", "联系人背景",
                    "业务关联度", "关联度理由", "信息来源"], "E2EFDA"),
    ("③ 开发信", ["邮件主题", "开发信正文", "生成备注"], "FFF2CC"),
]
WIDTHS = [6, 12, 12, 14, 20, 20, 26, 20, 22, 26, 28, 28,
          40, 16, 20, 26, 10, 30, 30, 34, 60, 18]
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

wb = Workbook()
ws = wb.active
ws.title = "客户表"
col = 1
for gname, fields, color in GROUPS:
    start = col
    for f in fields:
        c1 = ws.cell(row=1, column=col, value=gname if col == start else None)
        c2 = ws.cell(row=2, column=col, value=f)
        fill = PatternFill("solid", fgColor=color)
        for c in (c1, c2):
            c.fill = fill
            c.border = THIN
            c.alignment = Alignment(horizontal="center", vertical="center")
        c1.font = Font(name="Arial", bold=True, size=11)
        c2.font = Font(name="Arial", bold=True, size=10)
        col += 1
    if col - start > 1:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col - 1)

for i, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "D3"

ws2 = wb.create_sheet("使用说明")
notes = [
    "使用说明",
    "1. 本表由AI自动填写，业务员只需核对和复制开发信。",
    "2. 状态：⬜待处理 → 🔍已背调 → ✅已生成 → 📤已发送（最后一步业务员手动改）。",
    "3. 发送前务必人工核对邮箱、姓名拼写。",
    "4. 展会备注来自「交谈要点采集」文档和名片手写字，填得越具体开发信越个性化。",
]
for r, v in enumerate(notes, 1):
    ws2.cell(row=r, column=1, value=v).font = Font(name="Arial", size=10, bold=(r == 1))
ws2.column_dimensions["A"].width = 100

wb.save(OUT)
print(f"模板已生成: {OUT}")

# 嵌入名片缩略图的标准做法（供参考）：
# from openpyxl.drawing.image import Image as XLImage
# from PIL import Image
# img = Image.open(照片路径).convert("RGB")
# r = 200 / img.width                      # 宽度上限200px≈2.1英寸
# img = img.resize((200, int(img.height * r)))
# img.save(缩略图临时路径)
# xi = XLImage(缩略图临时路径); xi.anchor = f"L{行号}"; ws.add_image(xi)
# 数据行 row_dimensions[行号].height = 200
